"""Convertit l'export historique en classeur de saisie manuel simple."""

from __future__ import annotations

import argparse
import os
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PATH = PROJECT_ROOT / "data" / "postes_d3s.xlsx"

COLUMNS: list[tuple[str, tuple[str, ...]]] = [
    ("Actif ?", ("Actif ?", "poste_actif", "statut_poste")),
    ("Poste", ("Poste", "intitule_poste", "fonction")),
    ("Catégorie", ("Catégorie", "categorie_publication")),
    ("Établissement(s)", ("Établissement(s)", "etablissements")),
    ("Lieu(x) du poste", ("Lieu(x) du poste", "lieux_poste")),
    ("Ville", ("Ville", "ville_principale")),
    ("Département", ("Département", "departement")),
    ("Région", ("Région", "region")),
    ("FINESS", ("FINESS", "finess", "finess_et")),
    ("Date de parution", ("Date de parution", "date_jo")),
    ("Source", ("Source", "url_source")),
    ("Observations", ("Observations", "commentaire_correction")),
    ("Latitude", ("Latitude", "latitude")),
    ("Longitude", ("Longitude", "longitude")),
    ("Type de structure", ("Type de structure", "type_structure")),
    ("Type d'établissement", ("Type d'établissement", "type_etablissement")),
]

HEADER_NOTES = {
    "Actif ?": "Oui = visible sur la carte. Non = conservé dans Excel mais exclu de la carte.",
    "Poste": "Obligatoire. Intitulé court affiché dans la fiche.",
    "Catégorie": "Obligatoire. Choisir une valeur dans la liste.",
    "Établissement(s)": "Obligatoire. Nom du ou des établissements concernés.",
    "FINESS": "Numéro FINESS à 9 chiffres. Permet de calculer automatiquement les coordonnées.",
    "Latitude": "Facultatif si le FINESS est reconnu. Repli nécessaire sinon.",
    "Longitude": "Facultatif si le FINESS est reconnu. Repli nécessaire sinon.",
    "Type de structure": "Facultatif : le programme essaie de le déduire.",
    "Type d'établissement": "Facultatif : le programme essaie de le déduire.",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_PATH)
    return parser.parse_args()


def clean_value(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def select_value(row: pd.Series, candidates: tuple[str, ...]) -> Any:
    for candidate in candidates:
        if candidate not in row.index:
            continue
        value = clean_value(row[candidate])
        if value != "":
            if candidates[0] == "Actif ?" and candidate == "statut_poste":
                return "Non" if str(value).strip().casefold() == "retiré" else "Oui"
            if candidates[0] == "Date de parution":
                parsed = pd.to_datetime(value, errors="coerce")
                if not pd.isna(parsed):
                    return parsed.to_pydatetime()
            return value
    return "Oui" if candidates[0] == "Actif ?" else ""


def convert(input_path: Path, output_path: Path) -> None:
    with pd.ExcelFile(input_path, engine="openpyxl") as excel:
        available_sheets = excel.sheet_names
    source_sheet = "Postes_enrichis" if "Postes_enrichis" in available_sheets else "Postes"
    if source_sheet not in available_sheets:
        raise ValueError("Aucun onglet 'Postes_enrichis' ou 'Postes' trouvé.")
    source = pd.read_excel(input_path, sheet_name=source_sheet, engine="openpyxl")

    rows = [
        [select_value(row, candidates) for _, candidates in COLUMNS]
        for _, row in source.iterrows()
    ]
    workbook = build_workbook(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(suffix=".xlsx", dir=output_path.parent, delete=False) as handle:
        temporary_path = Path(handle.name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
        os.chmod(output_path, 0o644)
    finally:
        temporary_path.unlink(missing_ok=True)
    print(f"Classeur simplifié créé: {output_path} ({len(rows)} lignes, {len(COLUMNS)} colonnes)")


def build_workbook(rows: list[list[Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Postes"
    headers = [name for name, _ in COLUMNS]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    inactive_fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell.value in HEADER_NOTES:
            cell.comment = Comment(HEADER_NOTES[cell.value], "Projet D3S")

    widths = [11, 28, 34, 48, 30, 23, 20, 23, 14, 16, 36, 38, 12, 12, 22, 32]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 34
    sheet.sheet_view.showGridLines = False

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[8].number_format = "@"
        row[9].number_format = "dd/mm/yyyy"
        row[12].number_format = "0.000000"
        row[13].number_format = "0.000000"

    if rows:
        reference = f"A1:P{len(rows) + 1}"
        table = Table(displayName="TableauPostes", ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        sheet.add_table(table)
        sheet.conditional_formatting.add(
            f"A2:P{len(rows) + 1}",
            FormulaRule(formula=['LOWER($A2)="non"'], fill=inactive_fill),
        )

    add_list_validation(sheet, "A2:A10000", '"Oui,Non"')
    add_list_validation(
        sheet,
        "C2:C10000",
        '"Chef d\'établissement,Directeur adjoint,Poste réservé aux élèves directeurs,Autre / non précisé"',
    )
    add_list_validation(sheet, "O2:O10000", '"Sanitaire,Médico-social,Social,Non précisé"')
    return workbook


def add_list_validation(sheet: Any, cell_range: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choisissez une valeur proposée dans la liste."
    validation.errorTitle = "Valeur non reconnue"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def main() -> int:
    args = parse_args()
    convert(args.input.resolve(), args.output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
